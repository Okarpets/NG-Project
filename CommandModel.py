import requests
from requests.exceptions import Timeout
from requests.exceptions import ConnectionError
from openpyxl import load_workbook 
from bs4 import BeautifulSoup
import json


#COMAND 0 - HELP
def help():
    print(
    "This program was created as a website tester, below you can see a list of existing commands:\n\n" +
    "\t-help -- Write a manual for using the program\n" +
    "\t-exit -- Exits the program\n" +
    "\t-code_get <url> -- Returns the code from your get requests to the url\n" +
    "\t-code_post <url> <json_file name/without> -- Returns the code from your post requests to the url\n" +
    "\t-byid <url> <id> -- Returns the STATIC HTML element from your get requests to the url by id\n" +
    "\t-bytag <url> <tag> -- Returns the STATIC HTML element from your get requests to the url by tag\n" +
    "\t-create <json_name> -- Create scenario in document\n" +
    "\t-read <json_name> -- Read and process all scenario in document\n"
    )
    work()


#COMMAND 1 - RETURN GET CODE REQUESTS
def code_get(p_url):
    try:
        url = p_url
        try:
            site = requests.get(url)
        except Exception:
            print("\tInvalid values\n")
            work()
        code = site.status_code # CODE OF GET REQUEST
        print("\tReturned сode: " + str(code) + '\n')
    except ConnectionError: 
        print("\tConnection error\n")
    except Timeout:
        print("\tThe request timed out\n")
    file = 'TestResults.xlsx' #Create or open file with that name
    lst = load_workbook(file) #It is depend on save and close command
    xlsx = lst['CodeGetResults'] #List in 'TestResult.xlsx'
    xlsx.append([url,code]) #Adding that data in xlsx file
    lst.save(file) #REMEMBER THE WRITING RULES
    lst.close()
    work()


#COMMAND 2 - RETURN POST CODE REQUESTS
def code_post(orderarray):
    try:
        url = orderarray[1]
        post = orderarray[2]
        if post == "without":
            try:
                site = requests.post(url)
            except Exception:
                print("\tInvalid values\n")
                work()
            code = site.status_code # CODE OF POST REQUEST
            print("\tReturned сode: " + str(code) + '\n')
        else:
            try:
                file = open("{}.json".format(post), "r")
                u_data = json.load(file)
                site = requests.post(url, data=u_data)
            except Exception:
                print("\tInvalid values\n")
                work()
            code = site.status_code # CODE OF POST REQUEST
            print("\tReturned сode: " + str(code) + '\n')
    except ConnectionError: 
        print("\tConnection error\n")
    except Timeout:
        print("\tThe request timed out\n")

    file = 'TestResults.xlsx' #Create or open file with that name
    lst = load_workbook(file) #It is depend on save and close command
    xlsx = lst['CodePostResults'] #List in 'TestResult.xlsx'
    xlsx.append([url,code]) #Adding that data in xlsx file
    lst.save(file) #REMEMBER THE WRITING RULES
    lst.close()
    work()


#COMMAND 3 - ELEMENT BY ID
def byid(p_url, p_id):
    url = p_url
    site = requests.get(url)
    soup = BeautifulSoup(site.text, 'html.parser')
    u_id = soup.find_all(id='{0}'.format(p_id))
    print(u_id)
    if u_id != []:
        u_id = "The element IS on the page"
    else:
        u_id = "The element ISN'T on the page"
    print('\t' + u_id + '\n')
    file = 'TestResults.xlsx' #Create or open file with that name
    lst = load_workbook(file) #It is depend on save and close command
    xlsx = lst['ElementById'] #List in 'TestResult.xlsx'
    xlsx.append([p_id,url,u_id]) #Adding that data in xlsx file
    lst.save(file) #REMEMBER THE WRITING RULES
    lst.close()
    work()


#COMAND 4 - FORMATS .XLSX DOCUMENT
def excel():
    file = 'TestResults.xlsx'
    lst = load_workbook(file) #It is depend on save and close command
    for sheet_name in lst.sheetnames:
        sheet = lst[(sheet_name)]
        lst.remove(sheet)
    lst.create_sheet('ScenarioResults')
    lst.create_sheet('CodeGetResults') # CREATE LISTS
    lst.create_sheet('CodePostResults')
    lst.create_sheet('ElementById')
    lst.create_sheet('ElementByTag')
    lst['CodeGetResults'].append(["Url","Returned Code"])
    lst['CodePostResults'].append(["Url","Returned Code"])
    lst['ElementById'].append(["Given Id","Url","Result"])
    lst['ElementByTag'].append(["Given Tag","Url","Result"])
    lst['ScenarioResults'].append(["Url","Get requests", "Post requests", "Finded by id","Finded by tag"])
    lst.save(file) #REMEMBER THE WRITING RULES
    lst.close()

    
#COMMAND 5 - ELEMENT BY TAG
def bytag(p_url, p_tag):
    url = p_url
    site = requests.get(url)
    soup = BeautifulSoup(site.text, 'html.parser')
    u_tag = soup.find_all('{0}'.format(p_tag))
    if u_tag == []:
        u_tag = "Tag ISN'T HERE"
    print('\t' + str(u_tag) + '\n')
    if u_tag != []:
        u_tag = "Tag IS here"
    file = 'TestResults.xlsx' #Create or open file with that name
    lst = load_workbook(file) #It is depend on save and close command
    xlsx = lst['ElementByTag'] #List in 'TestResult.xlsx'
    xlsx.append([p_tag,url,u_tag]) #Adding that data in xlsx file
    lst.save(file) #REMEMBER THE WRITING RULES
    lst.close()
    work()


#COMMAND 6 - CREATE SCENARIO MODEL
def create(p_name):
    file = open(f"{p_name}.json", 'w')
    file.write('[''\n')
    while True:
        print(
        "\t\t\tNow pleaase write all command we have to do:\n" +
        "\t\tCreate scenario params. If you want skip a command enter 0, else 1\n\n"
        "\t\t\t\tPLEASE FILL IN THE FORM\n\n"
        )
        url = str(input('\tEnter url: '))
        get = "0" 
        post = "0" 
        htmlid = "0" 
        htmltag = "0" 
        command = "0"
        if input('\tDo you want do get requests?\n') == "1":
            get = "1"
        if input('\tDo you want do post requests?\n') == "1":
            post = "1"
        if input('\tDo you want search element by id?\n') == "1":
            htmlid = str(input('\tEnter id: '))
        if input('\tDo you want search element by tag?\n') == "1":
            htmltag = str(input('\tEnter tag: '))
        if get == "0" and post == "0" and htmlid == "0" and htmltag == "0":
            print("\t\tEmpty scenatio won't be added\n")
            create(p_name)
        data = {"url" : url, "get" : get, "post" : post, "htmlid" : htmlid, "htmltag" : htmltag}
        command = str(input("\n\t\tYou want create new case? If yes enter any button, else - \"F\"\n"))
        if command == 'F' or command == "f":
            json.dump(data, file)
            file.write('\n'']')
            file.close()
            break
        else:
            json.dump(data, file)
            file.write(',\n')
            

#COMAND 7 - SHOW ALL THE FILE OF SCENARIO
def show(p_name):
    with open("{0}.json".format(p_name), "r") as file:
        print("\t\tYour scenario:\n")
        for elem in file:
            print(elem)
    file.close()
    work()


#COMMAND 8 - READ SCENARIO MODEL
def read(p_name):
    file = open("{0}.json".format(p_name), "r")
    file = json.load(file) 
    for lines in file:
        url = lines['url']
        code_get = lines['get']
        if code_get == "1":
            site = requests.get(url)
            code_get = site.status_code
        else:
            code_get = "The operation was not requested"
        code_post = lines['post']
        if code_post == "1":
            site = requests.post(url)
            code_post = site.status_code
        else:
            code_post = "The operation was not requested"
        u_id = lines['htmlid']
        if u_id != "0":
            site = requests.get(url)
            soup = BeautifulSoup(site.text, 'html.parser')
            u_id = soup.find_all(id='{0}'.format(u_id))
            if u_id != []:
                u_id = "The element IS on the page"
            else:
                u_id = "The element ISN'T on the page"
        else:
            u_id = "The operation was not requested"
        u_tag = lines['htmltag']
        if u_tag != "0":
            site = requests.get(url)
            soup = BeautifulSoup(site.text, 'html.parser')
            u_tag = soup.find_all('{0}'.format(u_tag))
            if u_tag == []:
                u_tag = "Tag ISN'T HERE"
            if u_tag != []:
                u_tag = "Tag IS here"
        else:
            u_tag = "The operation was not requested"
        file = 'TestResults.xlsx' #Create or open file with that name
        lst = load_workbook(file) #It is depend on save and close command
        xlsx = lst['ScenarioResults'] #List in 'TestResult.xlsx'
        xlsx.append([url,code_get,code_post,u_id,u_tag]) #Adding that data in xlsx file
        lst.save(file) #REMEMBER THE WRITING RULES
        lst.close()
    print("\n\t\tAll the scenario was processed\n")
    work()

#COMMAND ANSWER - ERROR MESSAGES
def err():
    print("\t\tInvalid value\n\tTry again or use -help for help\n")
    work()


#COMMAND MAIN - START 
#STARTS A FUNCTION TREE
def work():
    order = str(input('\n'))
    orderarray = order.split(" ") # Вreaking the request into a command and a URL
    match orderarray[0]: # Getting the operation from the request
        case "-help":
                try:
                    help()
                except Exception:
                    err()
        case"-code_get":
                try:
                    p_url = orderarray[1]
                    code_get(p_url)
                except Exception:
                    err()
        case"-code_post":
                code_post(orderarray)
        case"-byid":
                try:
                    p_url = orderarray[1]
                    p_id = orderarray[2]
                    byid(p_url, p_id)
                except Exception:
                    err()
        case"-bytag":
                try:
                    p_url = orderarray[1]
                    p_tag = orderarray[2]
                    bytag(p_url, p_tag)
                except Exception:
                    err()
        case"-create":
                try:
                    p_name = orderarray[1]
                    create(p_name)
                except Exception:
                    err()
        case"-show":
                try:
                    p_name = orderarray[1]
                    show(p_name)
                except Exception:
                    err()
        case"-read":
                try:
                    p_name = orderarray[1]
                    read(p_name)
                except Exception:
                    err()
        case"-exit":
                    print("Exit the program")
                    return 0
        case _:
                print("\tError\n" + "\tWe're sorry, but this command doesn't exist, please use -help")
                work()


#MAIN !!!
print("\tWelcome, The program is running, enter the command or -help for manual to program\n")
lst = load_workbook('TestResults.xlsx')
if "ScenarioResults" != lst.sheetnames[0]:
    excel()

work()