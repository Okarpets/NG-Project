from openpyxl import load_workbook

file = 'TestResult.xlsx' #Create or open file with that name
lst = load_workbook(file) #It is depend on save and close command
ws = lst.create_sheet('Results') # CREATE LIST "Results"
xlsx = lst['Results'] #List in 'TestResult.xlsx'
xlsx.append(["Expected","Real Value","Result"]) #Adding that data in xlsx file
lst.save(file) #REMEMBER THE WRITING RULES
lst.close()
