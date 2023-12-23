from requests.exceptions import ConnectionError
from requests.exceptions import Timeout
from openpyxl import load_workbook 
from selenium import webdriver
from bs4 import BeautifulSoup
import xlsxwriter
import requests
import json
import os

OpusFile = "0"
driver=webdriver.Chrome()


#COMMAND USER : 0 - HELP (MAUNAL)
def help():
    print(
    "This program was created as a website tester, below you can see a list of existing commands:\n\n" +
    "\t-help -- Write a manual for using the program\n" +
    "\t-exit -- End program execution\n" +
    "\t-replace <new-xlsx-file name> -- Allows you to change the current Excel file for writing\n" +
    "\t-code_get <url> <json_file OR \"TC\" OR nothing> -- Returns the code from your get requests to the url\n" +
    "\t\t\"TC\" : headers, content, text -- typical commands to return a request (NOT WRITTEN IN EXCEL)\n" +
    "\t-code_post <url> <json_file OR nothing> -- Returns the code from your post requests to the url\n" +
    "\t-byid <url> <id> -- Returns the STATIC HTML element from your get requests to the url by id\n" +
    "\t-bytag <url> <tag> -- Returns the STATIC HTML element from your get requests to the url by tag\n" +
    "\t-create <json_name> -- Create scenario in a document\n" +
    "\t-process <json_name> <id> -- Process all scenario in a document OR one scenario in the file by scenario id\n" +
    "\t-show <json_name> <id> -- Show you all scenario in the file OR one scenario in the file by secanrio id\n" +
    "\t\tIf you don't enter id -show test show you all scenarios"
    )
    work()

#COMMAND USER : 1 - RETURN GET CODE REQUESTS (OR ANSWER AFTER PROCESS TYPICAL COMMANDS)
def code_get(orderarray):
    p_url = orderarray[1]
    try:
        if len(orderarray) == 2:
            try:
                site = requests.get(p_url)
            except Exception:
                print("\t\t\tInvalid values\n")
                work()
            order = site.status_code # CODE OF GET REQUEST
            print("\t\tReturned сode: " + str(order) + '\n')
        if len(orderarray) == 3:
            site = requests.get(p_url)   
            match orderarray[2]: # Getting the operation from the request
                case "headers":
                    try:
                        print(site.headers)
                        work()
                    except Exception:
                        err()
                case "text":
                    try:
                        print(site.text)
                        work()
                    except Exception:
                        err()
                case "content":
                    try:
                        print(site.content)
                        work()
                    except Exception:
                        err()
                case _:
                    try:
                        file = open("{}.json".format(orderarray[2]), "r")
                        u_data = json.load(file)
                        site = requests.get(p_url, params=json.dumps(u_data))
                        order = ("The request was processed correctly")
                        print("\t\tReturned: " + order + '\n')
                    except Exception:
                        order = "Uncorrectly requests"
                        print("\t\t{0}\n".format(order))
                        work()
    except ConnectionError: 
        print("\t\t\tConnection error\n")
    except Timeout:
        print("\t\t\tThe request timed out\n")
    file = OpusFile #Create or open file with that name
    lst = load_workbook(file) #It is depend on save and close command
    xlsx = lst['CodeGetResults'] #List in 'TestResult.xlsx'
    xlsx.append([p_url,order]) #Adding that data in xlsx file
    lst.save(file) #REMEMBER THE WRITING RULES
    lst.close()
    work()

#COMMAND USER : 2 - RETURN POST CODE REQUESTS
def code_post(orderarray):
    try:
        url = orderarray[1]
        if len(orderarray) == 2:
            try:
                site = requests.post(url)
            except Exception:
                print("\t\t\tInvalid values\n")
                work()
            order = site.status_code # CODE OF POST REQUEST
            print("\t\tReturned сode: " + str(order) + '\n')
        if len(orderarray) == 3:
            post = orderarray[2]
            try:
                file = open("{}.json".format(post), "r")
                u_data = json.load(file)
                site = requests.post(url, data=json.dumps(u_data))
            except Exception:
                print("\t\tUncorrectly requests\n")
                work()
            order = site.status_code # CODE OF POST REQUEST
            print("\t\tReturned сode: " + str(order) + '\n')
    except ConnectionError: 
        print("\t\tConnection error\n")
    except Timeout:
        print("\t\tThe request timed out\n")
    file = OpusFile #Create or open file with that name
    lst = load_workbook(file) #It is depend on save and close command
    xlsx = lst['CodePostResults'] #List in 'TestResult.xlsx'
    xlsx.append([url,order]) #Adding that data in xlsx file
    lst.save(file) #REMEMBER THE WRITING RULES
    lst.close()
    work()


#COMMAND USER : 3 - REPLACE MAIN EXCEL FILE
def repl(p_name):
    file = "{0}.xlsx".format(p_name)
    global OpusFile
    old_file = OpusFile
    OpusFile = file
    print(f"\n\t\tThe Excel file for recording was successfully changed from {old_file} to {OpusFile}\n")
    work()


#COMMAND USER : 4 - RETURN HTML ELEMENT BY ID
def byid(p_url, p_id):
    url = p_url
    driver.get(url)
    soup = BeautifulSoup(driver.page_source, 'lxml')
    u_id = soup.find_all(id='{0}'.format(p_id))
    if u_id == []:
        u_id = "Id ISN'T on the page"
    else:
        u_id = "Id IS on the page"
    print('\t' + u_id + '\n')
    file = OpusFile #Create or open file with that name
    lst = load_workbook(file) #It is depend on save and close command
    xlsx = lst['ElementById'] #List in 'TestResult.xlsx'
    xlsx.append([p_id,url,u_id]) #Adding that data in xlsx file
    lst.save(file) #REMEMBER THE WRITING RULES
    lst.close()
    work()


#COMMAND USER : 5 - RETURN HTML ELEMENT BY TAG
def bytag(p_url, p_tag):
    url = p_url
    driver.get(url)
    soup = BeautifulSoup(driver.page_source, 'lxml')
    u_tag = soup.find_all('{0}'.format(p_tag))
    if u_tag == []:
        u_tag = "Tag ISN'T on the page"
    else:
        u_tag = "Tag IS on the page"
    print('\t' + u_tag + '\n')
    file = OpusFile #Create or open file with that name
    lst = load_workbook(file) #It is depend on save and close command
    xlsx = lst['ElementByTag'] #List in 'TestResult.xlsx'
    xlsx.append([p_tag,url,u_tag]) #Adding that data in xlsx file
    lst.save(file) #REMEMBER THE WRITING RULES
    lst.close()
    work()
    

#COMMAND USER : 6 - CREATE SCENARIO TO JSON FILE
def create(p_name):
    file = open(f"{p_name}.json", 'w')
    file.write('[''\n')
    scen_id = 1
    while True:
        print("\tPlease enter url for {0}th scenario: ".format(scen_id))
        url = str(input())
        get = "0" 
        post = "0" 
        htmlid = "0" 
        htmltag = "0"
        get_params = "0"
        post_data = "0"
        oper = "0"
        oper_endp = "0"
        def under_create():
            print(
            "\t\t\tCommand list:\n" +
            "\t\t1 - Get requests\n" +
            "\t\t2 - Post requests\n" +
            "\t\t3 - Element by id\n" +
            "\t\t4 - Element by tag\n\n" +
            "\t\tE - Exit\n" +
            "\t\tC - Go to next scenario\n"
            )
            order = str(input("\t\tSelect a command from \"Command list\"\n"))
            return order
        while oper != "end":
            match under_create(): # Getting the operation from the request
                case "1":
                    par_order = str(input("\t\tDo you wanna add params? (Y/ )\n"))
                    if par_order == 'Y' or par_order == 'y':
                        get_params = str(input("\t\tEnter name of the json-file with params\n"))
                        get = "2"
                    else:
                        get = "1"
                case "2":       
                    par_order = str(input("\t\tDo you wanna to add data? (Y/ )\n"))
                    if par_order == 'Y' or par_order == 'y':
                        post_data = str(input("\t\tEnter name of the json-file with data\n"))
                        post = "2"
                    else:
                        post = "1"
                case "3":                   
                    htmlid = str(input("\t\tEnter HTML elemnent id\n"))                       
                case "4":                            
                    htmltag = str(input("\t\tEnter HTML elemnent tag\n"))
                case "E":
                    json.dump(data, file)
                    file.write('\n'']')
                    file.close()
                    oper = "end"      
                    oper_endp = "end"
                case "C":
                    if get == "0" and post == "0" and htmlid == "0" and htmltag == "0":
                        print("\tThe scenario will not be added because it's empty\n")
                    else:
                        json.dump(data, file)
                        file.write(',\n')
                        scen_id += 1
                        oper = "end"
                case _:
                    print("\tThis command doesn't exist, use \"Command list\"\n")
            data = {"id" : scen_id, "url" : url, "get" : get, "params" : get_params, "post" : post, "data" : post_data, "htmlid" : htmlid, "htmltag" : htmltag}
        if oper_endp == "end":
            break
            
#COMMAND USER : 7 - COMMAND TO CREATE REQUESTS FOR SCENARIO PROCESSING IN AN EXCEL FILE
def process(p_name, p_id):
    if p_id == "0":
        file = open("{0}.json".format(p_name), "r")
        file = json.load(file) 
        for lines in file:
            handle(lines, p_name)
        print("\n\t\tAll the scenario was processed\n")
    else:  
        with open("{0}.json".format(p_name), "r") as file:
            js_file = json.load(file)
            file.close()
            for elem in js_file:
                if str(elem['id']) == str(p_id):
                    handle(elem, p_name)
        print("\n\t\tScenario was processed\n")
    work()


#COMMAND USER : 8 - SHOW ALL THE FILE OF SCENARIO
def show(p_name, p_id):
    with open("{0}.json".format(p_name), "r") as file:
        if p_id == "0":
            print("\t\tYour scenario:\n")
            for lines in file:
                print(lines)
            file.close()
            work()
        js_file = json.load(file)
        file.close()
        print("\t\tYour scenario:\n")
        for lines in js_file:
            if str(lines['id']) == str(p_id):
                print(lines)
                work()
        print("\n\tScenario with that id wasn't found or it doesn't exist\n")
        work()


#TECHNICAL COMMANDS NOT USER APPLICABLE
############################################################################
    
#COMMAND ANSWER : 1 - FORMATS .XLSX DOCUMENT
def excel():
    file = str(input("\t\t\tPlease enter the Excel file name for the entry\n"))
    file_list = os.listdir()
    file = "{0}.xlsx".format(file)
    global OpusFile
    OpusFile = file
    if file in file_list:
        lst = load_workbook(file)
        if "ScenarioResults" != lst.sheetnames[0]:
            formating(file)
            print("\t\t\t\tFile found and formatted\n")
        else:
            print("\t\t\t\t\tFile found\n")
    else:
        order = str(input("\t\tThat file doesn't exist, do you wanna create it? (Y/ )\n"))
        if order == 'Y' or order == 'y':
            workbook = xlsxwriter.Workbook(file)
            worksheet = workbook.add_worksheet()
            workbook.close()
            formating(file)
            print("\t\t\tThe file was created successfully\n")
        else:
            excel()


#COMMAND ANSWER : 2 - ERROR MESSAGES
def err():
    print("\t\tInvalid value\n\tTry again or use -help for help\n")
    work()


#COMMAND ANSWER : 3 - PROCESS SCENARIOS
def handle(lines, p_name):
        scen_id = lines['id']
        url = lines['url']
        code_get = lines['get']
        get_params = lines['params']
        code_post = lines['post']
        post_data = lines['data']
        u_id = lines['htmlid']
        u_tag = lines['htmltag']

        #GET PART
        if code_get == "1":
            site = requests.get(url)
            get_order = site.status_code
        if code_get == "2":
            try:
                file = open("{}.json".format(get_params), "r")
                u_data = json.load(file)
                site = requests.get(url, params=json.dumps(u_data))
                get_order = site.status_code
            except Exception:
                get_order = "Uncorrectly requests"
        else:
            get_order = "The operation wasn't requested"

        #POST PART
        if code_post == "1":
            site = requests.post(url)
            post_order = site.status_code
        if code_post == "2":
            try:
                file = open("{}.json".format(post_data), "r")
                u_data = json.load(file)
                site = requests.post(url, data=json.dumps(u_data))
                post_order = site.status_code
            except Exception:
                post_order = "Uncorrectly requests"
        else:
            post_order = "The operation wasn't requested"

        #ID PART
        if u_id != "0":
            driver.get(url)
            soup = BeautifulSoup(driver.page_source, 'lxml')
            u_id = soup.find_all(id='{0}'.format(u_id))
            if u_id == []:
                u_id = "Id ISN'T on the page"
            else:
                u_id = "Id IS on the page"
        else:
            u_id = "The operation wasn't requested"
        
        #TAG PART
        if u_tag != "0":
            driver.get(url)
            soup = BeautifulSoup(driver.page_source, 'lxml')
            u_tag = soup.find_all('{0}'.format(u_tag))
            if u_tag == []:
                u_tag = "Tag ISN'T on the page"
            else:
                u_tag = "Tag IS on the page"
        else:
            u_tag = "The operation wasn't requested"
        file = OpusFile #Create or open file with that name
        lst = load_workbook(file) #It is depend on save and close command
        xlsx = lst['ScenarioResults'] #List in 'TestResult.xlsx'
        xlsx.append([scen_id,url,get_order,post_order,u_id,u_tag,p_name]) #Adding that data in xlsx file
        lst.save(file) #REMEMBER THE WRITING RULES
        lst.close()


#COMMAND ANSWER : 4 - FORMANTS AN EXCEL FILE 
def formating(file):
    lst = load_workbook(file) #It is depend on save and close command
    for sheet_name in lst.sheetnames:
        sheet = lst[(sheet_name)]
        lst.remove(sheet)
        lst.create_sheet('ScenarioResults')
        lst.create_sheet('CodeGetResults') # CREATE LISTS
        lst.create_sheet('CodePostResults')
        lst.create_sheet('ElementById')
        lst.create_sheet('ElementByTag')
        lst['CodeGetResults'].append(["Url","Order"])    
        lst['CodePostResults'].append(["Url","Order"])
        lst['ElementById'].append(["Given Id","Url","Result"])    
        lst['ElementByTag'].append(["Given Tag","Url","Result"])
        lst['ScenarioResults'].append(["Scenario ID","Url","Get requests", "Post requests", "Find-order by id","Find-order by tag", "Scenario-file name"])
        lst.save(file) #REMEMBER THE WRITING RULES
        lst.close()

############################################################################
    

#COMMAND MAIN - START 
#STARTS A FUNCTION TREE
def work():
    order = str(input('\n'))
    orderarray = order.split(" ") # Вreaking the request into a command and a URL
    match orderarray[0]: # Getting the operation from the request
        case"-help":
                try:
                    help()
                except Exception:
                    err()
        case"-exit":
                print("\n\t\tEnd program execution")
                return 0
        case"-replace":
                try:
                    p_name = orderarray[1]
                    repl(p_name)
                except Exception:
                    err()
        case"-code_get":
                code_get(orderarray)
        case"-code_post":
                code_post(orderarray)
        case"-byid":
                try:
                    p_url = orderarray[1]
                    p_id = orderarray[2]
                    byid(p_url, p_id)
                except Exception:
                    err()
        case"-bytag":
                try:
                    p_url = orderarray[1]
                    p_tag = orderarray[2]
                    bytag(p_url, p_tag)
                except Exception:
                    err()
        case"-create":
                try:
                    p_name = orderarray[1]
                    create(p_name)
                except Exception:
                    err()
        case"-process":
                try:
                    if len(orderarray) == 2:
                        p_name = orderarray[1]
                        process(p_name, "0")
                    if len(orderarray) == 3:
                        p_name = orderarray[1]
                        p_id = orderarray[2]
                        process(p_name, p_id)
                except Exception:
                    err()
        case"-show":
                try:
                    if len(orderarray) == 2:
                        p_name = orderarray[1]
                        show(p_name, "0")
                    if len(orderarray) == 3:
                        p_name = orderarray[1]
                        p_id = orderarray[2]
                        show(p_name, p_id)
                except Exception:
                    err()
        case _:
                err()
                work()


#MAIN !!!
print("\n\t\tWelcome! The program has started. Type the command or -help for programming guidance or use right now\n")
excel()
work()

